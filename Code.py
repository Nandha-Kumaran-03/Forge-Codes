import time
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC


# Load data from Excel
wb = load_workbook('D:\Summer 23\SPI\Dolabs\Data.xlsx')
sheet = wb.active


# Setup WebDriver
driver = webdriver.Chrome()

wait = WebDriverWait(driver, 10)

# Open WhatsApp Web
driver.get("https://web.whatsapp.com")
print("Scan the QR code with your phone to log in to WhatsApp Web")
time.sleep(15)  # Wait for user to scan QR code

# Iterate through each row in the Excel sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    Name, Group_Number, Teammates, Phone_Number = row

    # Format message
    message = f"Hi {Name},\nyou have been allotted in group {Group_Number},\nyour teammates are {Teammates}."

    # Open chat
    driver.get(f"https://web.whatsapp.com/send?phone={Phone_Number}")

    # Wait for the chat to load
    time.sleep(10)

    # Locate the message box
    message_box = wait.until(EC.presence_of_element_located(by="xpath",value=
        '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]'))
    message_box.send_keys(message)

    # Send message
    message_box.send_keys(message + Keys.ENTER)

    # Wait for the message to be sent
    time.sleep(3)

# Close the WebDriver
driver.quit()
